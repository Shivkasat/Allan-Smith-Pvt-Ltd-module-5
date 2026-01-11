import tkinter as tk
from tkinter import filedialog, messagebox
import os
import pandas as pd
import numpy as np
from fpdf import FPDF
from openpyxl import load_workbook
from tkcalendar import DateEntry
import webbrowser
import matplotlib.pyplot as plt
import glob



def process_file(file_path, company_name, equipment_name, feed_rate, date_of_measurement, user_inp, no_of_pier, radar_positions):
    angle_increment = 360 / int(user_inp)


    try:
        data = pd.read_excel(file_path, sheet_name=0)
        filtered_data = data[data['CHAIRPAD NO'].apply(lambda x: str(x).isnumeric())]

        if filtered_data.empty:
            messagebox.showerror("Data Error", "No valid CHAIRPAD NO data found in the Excel file.")
            return

        filtered_data.columns = [chr(65 + i) for i in range(len(filtered_data.columns))]
    except Exception as e:
        messagebox.showerror("File Error", f"Could not read Excel file: {str(e)}")
        return


    position = list(range(1, int(user_inp) + 1)) + [1]
    measurement = [i * angle_increment for i in range(int(user_inp))] + [360]


    try:
        distance_row = data.iloc[68, 1:].values
        cumulative_distance_row = data.iloc[69, 1:].values
        Diff_temp = data.iloc[70, 1:].values - data.iloc[71, 1:].values
        Min_temp = data.iloc[70, 1:].values
        Max_temp = data.iloc[71, 1:].values
        AVG_temp = data.iloc[72, 1:].values
    except IndexError:
        messagebox.showerror("Data Error", "Excel file does not have the required rows (68-72). Please check your file format.")
        return


    summary_data = []
    Temp_data = []
    all_sheet_data = {}


    output_file = "processed_data_with_summary.xlsx"

    num_columns = len(filtered_data.columns[1:])

    if num_columns == 0:
        messagebox.showerror("Data Error", "No data columns found after filtering.")
        return

    for i, col in enumerate(filtered_data.columns[1:]):
        try:
            col_data = filtered_data[col].tolist()

            if len(col_data) == 0:
                print(f"[Warning] Column {col} is empty, skipping...")
                continue

            data_measured = col_data + [col_data[0]]
            data_measured = pd.to_numeric(data_measured, errors='coerce')

            data_measured_clean = [x for x in data_measured if not np.isnan(x)]

            if len(data_measured_clean) == 0:
                print(f"[Warning] Column {col} has no valid numeric data, skipping...")
                continue

            max_measured = np.nanmax(data_measured)
            shell_run_out = [max_measured - value if not np.isnan(value) else 0 for value in data_measured]


            max_length = max(len(position), len(measurement), len(data_measured), len(shell_run_out))

            pos = position[:max_length] + [np.nan] * (max_length - len(position))
            meas = measurement[:max_length] + [np.nan] * (max_length - len(measurement))
            data_measured_padded = list(data_measured) + [np.nan] * (max_length - len(data_measured))
            shell_run_out_padded = shell_run_out + [np.nan] * (max_length - len(shell_run_out))


            sheet_data = pd.DataFrame({
                'Position': pos,
                'Measurement': meas,
                'Data Measured': data_measured_padded,
                'Distortion': data_measured_padded,
                'Run Out': shell_run_out_padded
            })


            sheet_data['AA'] = sheet_data['Measurement'] / 180 * 3.14
            sheet_data['AB'] = np.cos(sheet_data['AA']) * sheet_data['Run Out']
            sheet_data['AC'] = np.sin(sheet_data['AA']) * sheet_data['Run Out']


            SUM_AB = sheet_data['AB'][:-1].sum()
            SUM_AC = sheet_data['AC'][:-1].sum()


            XX = 2 / int(user_inp) * SUM_AB
            YY = 2 / int(user_inp) * SUM_AC
            ZZ = np.sqrt(XX ** 2 + YY ** 2)


            Angle_of_Occurrence = np.arccos(XX / ZZ) * 180 / 3.14 if ZZ != 0 else 0
            if YY < 0:
                Angle_of_Occurrence = 360 - Angle_of_Occurrence


            sheet_data['AD'] = (Angle_of_Occurrence - sheet_data['Measurement']) / 180 * 3.14
            sheet_data['AE'] = np.cos(sheet_data['AD'])
            sheet_data['AF'] = ZZ * sheet_data['AE']
            sheet_data['AG'] = sheet_data['Run Out'] - sheet_data['AF']
            AVG_AG = sheet_data['AG'][:-1].mean()
            sheet_data['AH'] = sheet_data['AG'] - AVG_AG
            sheet_data['Distortion'] = sheet_data['AH']
            sheet_data['AI'] = sheet_data['AF'] + AVG_AG


            summary_data.append({
                'Position': i + 1,
                'X': XX,
                'Y': YY,
                'Eccentricity (mm)': ZZ,
                'Phase Angle': Angle_of_Occurrence,
                'Runout': np.nanmax(shell_run_out),
                'Local Shell Deformation': AVG_AG,
                'Distance': distance_row[i] if i < len(distance_row) else np.nan,
                'Cumulative Distance': cumulative_distance_row[i] if i < len(cumulative_distance_row) else np.nan,
            })

            Temp_data.append({
                'Position': i + 1,
                'Diff': Diff_temp[i] if i < len(Diff_temp) else np.nan,
                'Min': Min_temp[i] if i < len(Min_temp) else np.nan,
                'Max': Max_temp[i] if i < len(Max_temp) else np.nan,
                'AVG': AVG_temp[i] if i < len(AVG_temp) else np.nan,
            })


            all_sheet_data[f"Sheet_{i+1}"] = sheet_data

        except Exception as e:
            print(f"[Error] Processing column {col}: {str(e)}")
            import traceback
            traceback.print_exc()
            continue


    if len(all_sheet_data) == 0:
        messagebox.showerror("Processing Error", "No valid data could be processed. Please check your Excel file format.")
        return


    summary_df = pd.DataFrame(summary_data)
    Temp_df = pd.DataFrame(Temp_data)


    try:
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
            for sheet_name, sheet_df in all_sheet_data.items():
                sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

            Temp_df.to_excel(writer, sheet_name='Temp', index=False)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

        print(f"[Success] Excel file created: {output_file}")

    except Exception as e:
        messagebox.showerror("Excel Error", f"Failed to create Excel file: {str(e)}")
        return

    messagebox.showinfo("Success", f"File processed and saved as {output_file}")

    try:
        generate_pdf(
            output_file,
            "processed_report.pdf",
            company_name,
            equipment_name,
            feed_rate,
            date_of_measurement,
            no_of_pier,
            radar_positions
        )
    except Exception as e:
        messagebox.showerror("PDF Error", f"Failed to generate PDF: {str(e)}")
        import traceback
        traceback.print_exc()



# Function to create a radar chart - USER DEFINED POSITION LIMIT
def create_radar_chart(ax, Run_out, title, max_positions):
    """
    Create radar chart showing only Run Out data with plain line
    Shows only specified number of positions
    Run_out: Actual runout data from Run Out column
    max_positions: Maximum number of positions to display
    """
    try:
        # Convert to list if pandas Series
        if isinstance(Run_out, pd.Series):
            Run_out = Run_out.tolist()

        # Remove NaN values
        Run_out = [x for x in Run_out if not pd.isna(x)]

        # Remove last duplicate point
        Run_out = Run_out[:-1]


        # Limit to user-specified number of positions
        if len(Run_out) > max_positions:
            Run_out = Run_out[:max_positions]
            print(f"[Info] Limiting radar chart to first {max_positions} positions")


        # Flip data
        Run_out_flipped = Run_out[1:] + Run_out[:1]


        # Number of variables
        num_vars = len(Run_out_flipped)


        if num_vars < 3:
            print("[Warning] Need at least 3 data points for radar chart")
            return


        # Create angles for the radar chart
        angles = np.linspace(0, 2 * np.pi, num_vars, endpoint=False).tolist()


        # Close the plot
        Run_out_values = Run_out_flipped + [Run_out_flipped[0]]
        angles_plot = angles + [angles[0]]


        # Draw radar chart
        ax.set_theta_offset(np.pi / 2)
        ax.set_theta_direction(-1)

        # Set position labels
        position_labels = [f'{i+1}' for i in range(num_vars)]
        plt.xticks(angles, position_labels, color='black', size=10, weight='bold')

        # Set proper y-axis limits based on data
        y_min = min(Run_out_flipped) - 10
        y_max = max(Run_out_flipped) + 10
        ax.set_ylim(y_min, y_max)

        # Add radial grid lines with labels
        y_ticks = np.linspace(y_min, y_max, 5)
        ax.set_yticks(y_ticks)
        ax.set_yticklabels([f'{y:.1f}' for y in y_ticks], size=8)

        # Plot ONLY plain line - NO fill, NO markers
        ax.plot(angles_plot, Run_out_values, color='blue', linewidth=3, label='Run Out')


        # Styling
        ax.set_title(title, size=14, y=1.1, weight='bold', pad=20)
        ax.legend(loc='upper right', bbox_to_anchor=(1.2, 1.1), fontsize=11, frameon=True, shadow=True)

        # Add professional grid
        ax.grid(True, linestyle='--', alpha=0.7, linewidth=1)

        print(f"[Success] Radar chart created with {num_vars} positions")

    except Exception as e:
        print(f"[Error] Creating radar chart: {e}")
        import traceback
        traceback.print_exc()

def generate_pdf(excel_path, pdf_path, company_name, equipment_name,
                 feed_rate, date_of_measurement, no_of_pier, radar_positions):

    import os, glob, webbrowser
    import pandas as pd
    import numpy as np
    import matplotlib.pyplot as plt
    from openpyxl import load_workbook
    from fpdf import FPDF
    from tkinter import messagebox

    # ================= FILE CHECK =================
    if not os.path.exists(excel_path):
        messagebox.showerror("Error", f"Excel file not found: {excel_path}")
        return

    try:
        wb = load_workbook(excel_path, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
    except Exception as e:
        messagebox.showerror("Error", f"Could not read Excel file: {str(e)}")
        return

    # ================= PDF CLASS =================
    class PDF(FPDF):
        def footer(self):
            self.set_y(-15)
            self.set_font('Arial', 'I', 10)
            self.set_text_color(255, 0, 0)
            self.cell(0, 10, 'Allan Smith Engineering Pvt. Ltd.', 0, 0, 'R')

    pdf = PDF()
    pdf.set_auto_page_break(auto=True, margin=15)

    filtered_sheet_names = [
        name for name in sheet_names if name.lower() not in ['summary', 'temp']
    ]

    if not filtered_sheet_names:
        messagebox.showerror("Error", "No valid sheets found in Excel file")
        return

    # ================= MAIN LOOP =================
    for idx, sheet_name in enumerate(filtered_sheet_names):
        try:
            df = pd.read_excel(excel_path, sheet_name=sheet_name)
            pdf.add_page()

            # ================= LOGO =================
            if os.path.exists("companylogo.jpg"):
                pdf.image("companylogo.jpg", x=10, y=20, w=25)

            # ================= FIG IMAGE (EVERY PAGE) =================
            if os.path.exists("FIG.jpg"):
                pdf.image("FIG.jpg", x=110, y=240, w=100)

            # ================= HEADER =================
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(0, 10, "Roller shaft deflection Report", ln=True, align='C')

            pdf.set_font("Arial", 'B', 11)
            pdf.set_xy(40, 20); pdf.cell(0, 8, f"Company Name: {company_name}", ln=True)
            pdf.set_xy(40, 25); pdf.cell(0, 8, f"Equipment Name: {equipment_name}", ln=True)
            pdf.set_xy(40, 30); pdf.cell(0, 8, f"Capacity: {feed_rate}", ln=True)
            pdf.set_xy(40, 35); pdf.cell(0, 8, f"Date of Measurement: {date_of_measurement}", ln=True)
            pdf.set_xy(40, 40); pdf.cell(0, 8, "Method: Single Point", ln=True)
            pdf.set_xy(40, 45); pdf.cell(0, 8, f"No. of Pier: {no_of_pier}", ln=True)

            # ================= LINE GRAPH =================
            if 'Run Out' in df.columns and 'AI' in df.columns:
                try:
                    plt.figure(figsize=(7, 3.5))
                    plt.plot(df['Run Out'].dropna(), label='Actual', linewidth=2)
                    plt.plot(df['AI'].dropna(), label='Reference', linewidth=2)
                    plt.xlabel("Position")
                    plt.ylabel("Value (mm)")
                    plt.title("Roller shaft deflection linear Graph\n(During single revolution of Kiln)")
                    plt.legend()
                    plt.grid(True)

                    graph_path = f"temp_graph_{sheet_name}.png"
                    plt.savefig(graph_path, dpi=150)
                    plt.close()

                    if os.path.exists(graph_path):
                        pdf.image(graph_path, x=110, y=180, w=100)
                except Exception as e:
                    print("[Graph Error]", e)
            try:
                summary_data = pd.read_excel(excel_path, sheet_name='Summary')
                angle_of_occurrence_values = summary_data['Phase Angle'].dropna().tolist()
                runout_values = summary_data['Runout'].dropna().tolist()
                eccentricity_values = summary_data['Eccentricity (mm)'].dropna().tolist()

                angle_of_occurrence_value = angle_of_occurrence_values[idx] if idx < len(angle_of_occurrence_values) else "N/A"
                eccentricity_value = eccentricity_values[idx] if idx < len(eccentricity_values) else "N/A"
                runout_value = runout_values[idx] if idx < len(runout_values) else "N/A"
            except Exception as e:
                print(f"[Warning] Could not read summary data: {e}")
                angle_of_occurrence_value = "N/A"
                eccentricity_value = "N/A"
                runout_value = "N/A"

            # Result metrics on RIGHT SIDE
            pdf.set_font("Arial", style="B", size=13)
            pdf.set_xy(140, 55)
            pdf.cell(0, 10, f"Result:", align='R')
            pdf.set_font("Arial", size=13)
            pdf.set_xy(140, 62)
            pdf.cell(0, 10, f"Run out Range = {runout_value:.2f} mm" if isinstance(runout_value, (int, float)) else f"Run out Range = {runout_value} mm", ln=True, align='R')
            pdf.set_xy(140, 69)
            pdf.cell(0, 10, f"Angle of Occurrence = {angle_of_occurrence_value:.2f}°" if isinstance(angle_of_occurrence_value, (int, float)) else f"Angle of Occurrence = {angle_of_occurrence_value}°", ln=True, align='R')
            pdf.set_xy(140, 76)
            pdf.cell(0, 10, f"Eccentricity = {eccentricity_value:.2f} mm" if isinstance(eccentricity_value, (int, float)) else f"Eccentricity = {eccentricity_value} mm", ln=True, align='R')

            # ================= RADAR CHART =================
            if 'Run Out' in df.columns and len(df['Run Out'].dropna()) > 2:
                try:
                    fig = plt.figure(figsize=(4.5, 4.5))
                    ax = fig.add_subplot(111, polar=True)
                    create_radar_chart(ax, df['Run Out'].dropna(),
                                        "Roller Raceway eccentricity\n& deformation Polar Graph",
                                        radar_positions)

                    radar_path = f"temp_radar_{sheet_name}.png"
                    plt.savefig(radar_path, dpi=150, bbox_inches='tight')
                    plt.close()

                    if os.path.exists(radar_path):
                        pdf.image(radar_path, x=130, y=95, w=65)
                except Exception as e:
                    print("[Radar Error]", e)

            # ================= TABLE =================
            columns_to_print = [c for c in df.columns if c != 'Distortion'][:4]
            pdf.set_xy(10, 60)

            cell_width = 20
            wide_width = 28
            cell_h_data = 5
            cell_h_header = 12
            line_h = 4

            header_texts = {
                'Position': 'Position',
                'Measurement': 'Measurement\nAngle',
                'Data Measured': 'Data\nMeasured',
                'Run Out': 'S.R.\nRun Out'
            }

            def get_col_width(col):
                return wide_width if col.lower() in ["measurement", "data measured"] else cell_width

            def print_table_header():
                pdf.set_font("Arial", 'B', 8)
                for col in columns_to_print:
                    text = header_texts.get(col, col)
                    width = get_col_width(col)

                    x, y = pdf.get_x(), pdf.get_y()
                    lines = text.count("\n") + 1
                    y_offset = y + (cell_h_header - lines * line_h) / 2

                    pdf.rect(x, y, width, cell_h_header)
                    pdf.set_xy(x, y_offset)
                    pdf.multi_cell(width, line_h, text, border=0, align='C')
                    pdf.set_xy(x + width, y)
                pdf.ln(cell_h_header)

            print_table_header()
            pdf.set_font("Arial", '', 8)

            for _, row in df.iterrows():
                if pd.isna(row['Position']):
                    continue

                if pdf.get_y() > 270:
                    pdf.add_page()
                    pdf.set_xy(10, 20)
                    
                    print_table_header()
                    pdf.set_font("Arial", '', 8)

                for col in columns_to_print:
                    width = get_col_width(col)
                    val = row[col]

                    if pd.isna(val):
                        txt = "N/A"
                    elif col.lower() == "position":
                        txt = str(int(val))
                    else:
                        txt = f"{float(val):.2f}"

                    pdf.cell(width, cell_h_data, txt, border=1, align='C')
                pdf.ln(cell_h_data)

            # ================= IMAGE BELOW TABLE =================
            if os.path.exists("TUPDN.jpg"):
                y = pdf.get_y()
                if y > 220:
                    pdf.add_page()
                    y = 20
                pdf.image("TUPDN.jpg", x=70, y=y + 5, w=70)
                pdf.ln(60)

        except Exception as e:
            print(f"[Error] Sheet {sheet_name}:", e)

    # ================= SAVE PDF =================
    pdf.output(pdf_path)

    # ================= CLEANUP =================
    for f in glob.glob("temp_graph_*.png") + glob.glob("temp_radar_*.png"):
        try:
            os.remove(f)
        except:
            pass

    messagebox.showinfo("PDF Generated", f"PDF report saved as {pdf_path}")

    try:
        os.startfile(pdf_path)
    except:
        webbrowser.open(pdf_path)





def browse_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")])
    if file_path:
        entry_file.delete(0, tk.END)
        entry_file.insert(0, file_path)



def on_submit():
    file_path = entry_file.get()
    company_name = entry_company.get()
    equipment_name = entry_equipment.get()
    feed_rate = entry_feed.get()
    date_of_measurement = entry_date.get()
    positions = entry_positions.get()
    no_of_pier = entry_pier.get()
    radar_pos_input = entry_radar_positions.get()


    if not all([file_path, company_name, equipment_name, feed_rate, date_of_measurement, positions, no_of_pier, radar_pos_input]):
        messagebox.showerror("Input Error", "All fields are required.")
        return


    try:
        radar_positions = int(radar_pos_input)
        if radar_positions < 3:
            messagebox.showerror("Input Error", "Radar chart positions must be at least 3.")
            return
    except ValueError:
        messagebox.showerror("Input Error", "Radar chart positions must be a valid number.")
        return


    try:
        process_file(file_path, company_name, equipment_name, feed_rate, date_of_measurement, int(positions), no_of_pier, radar_positions)
    except Exception as e:
        messagebox.showerror("Processing Error", f"Error: {str(e)}\n\nPlease check the console for details.")
        import traceback
        traceback.print_exc()



def show_about():
    about_window = tk.Toplevel()
    about_window.title("About Developer")
    about_window.geometry("350x220")


    tk.Label(about_window, text="Roller shaft deflection Report Generator", font=('Arial', 12, 'bold')).pack(pady=10)
    tk.Label(about_window, text="Version 1.0", font=('Arial', 10)).pack()
    tk.Label(about_window, text="Developed by Shiv Sunil Kasat", font=('Arial', 12, 'bold')).pack(pady=10)


    email_label = tk.Label(about_window, text="shivkasat477@gmail.com", fg="blue", cursor="hand2", font=('Arial', 10, 'underline'))
    email_label.pack()
    email_label.bind("<Button-1>", lambda e: webbrowser.open("mailto:shivkasat477@gmail.com"))


    linkedin_label = tk.Label(about_window, text="linkedin.com/in/shiv-kasat-005b3b252", fg="blue", cursor="hand2", font=('Arial', 10, 'underline'))
    linkedin_label.pack(pady=5)
    linkedin_label.bind("<Button-1>", lambda e: webbrowser.open("https://linkedin.com/in/shiv-kasat-005b3b252"))


    tk.Label(about_window, text="© 2025 Shiv Kasat", font=('Arial', 9)).pack(pady=10)



def show_main_app():
    login_window.destroy()
    global entry_company, entry_equipment, entry_feed, entry_date, entry_positions, entry_pier, entry_file, entry_radar_positions
    root = tk.Tk()
    root.title("Axial Runout Report Generator - Single Point")
    root.geometry("600x500")
    root.resizable(False, False)

    menu_bar = tk.Menu(root)
    help_menu = tk.Menu(menu_bar, tearoff=0)
    help_menu.add_command(label="About Developer", command=show_about)
    menu_bar.add_cascade(label="Help", menu=help_menu)
    root.config(menu=menu_bar)

    title_label = tk.Label(root, text="Roller shaft deflection Report Generator", font=('Arial', 16, 'bold'), bg='#2c3e50', fg='white')
    title_label.grid(row=0, column=0, columnspan=3, sticky='ew', pady=(0, 20))

    tk.Label(root, text="Select Excel File:", font=('Arial', 10)).grid(row=1, column=0, sticky='w', padx=20, pady=8)
    entry_file = tk.Entry(root, width=45, font=('Arial', 10))
    entry_file.grid(row=1, column=1, padx=5, pady=8)
    tk.Button(root, text="Browse", command=browse_file, bg='#3498db', fg='white', font=('Arial', 9, 'bold')).grid(row=1, column=2, padx=20, pady=8)


    tk.Label(root, text="Company Name:", font=('Arial', 10)).grid(row=2, column=0, sticky='w', padx=20, pady=8)
    entry_company = tk.Entry(root, width=45, font=('Arial', 10))
    entry_company.grid(row=2, column=1, padx=5, pady=8, columnspan=2, sticky='w')


    tk.Label(root, text="Equipment Name:", font=('Arial', 10)).grid(row=3, column=0, sticky='w', padx=20, pady=8)
    entry_equipment = tk.Entry(root, width=45, font=('Arial', 10))
    entry_equipment.grid(row=3, column=1, padx=5, pady=8, columnspan=2, sticky='w')


    tk.Label(root, text="Capacity:", font=('Arial', 10)).grid(row=4, column=0, sticky='w', padx=20, pady=8)
    entry_feed = tk.Entry(root, width=45, font=('Arial', 10))
    entry_feed.grid(row=4, column=1, padx=5, pady=8, columnspan=2, sticky='w')


    tk.Label(root, text="Date of Measurement:", font=('Arial', 10)).grid(row=5, column=0, sticky='w', padx=20, pady=8)
    entry_date = DateEntry(root, width=42, font=('Arial', 10))
    entry_date.grid(row=5, column=1, padx=5, pady=8, columnspan=2, sticky='w')


    tk.Label(root, text="No. of Positions:", font=('Arial', 10)).grid(row=6, column=0, sticky='w', padx=20, pady=8)
    entry_positions = tk.Entry(root, width=45, font=('Arial', 10))
    entry_positions.grid(row=6, column=1, padx=5, pady=8, columnspan=2, sticky='w')


    tk.Label(root, text="No. of Pier:", font=('Arial', 10)).grid(row=7, column=0, sticky='w', padx=20, pady=8)
    entry_pier = tk.Entry(root, width=45, font=('Arial', 10))
    entry_pier.grid(row=7, column=1, padx=5, pady=8, columnspan=2, sticky='w')


    tk.Label(root, text="Radar Chart Positions:", font=('Arial', 10)).grid(row=8, column=0, sticky='w', padx=20, pady=8)
    entry_radar_positions = tk.Entry(root, width=45, font=('Arial', 10))
    entry_radar_positions.insert(0, "21")  # Default value
    entry_radar_positions.grid(row=8, column=1, padx=5, pady=8, columnspan=2, sticky='w')
    tk.Label(root, text="(number of positions to show on radar chart)", font=('Arial', 8), fg='gray').grid(row=9, column=1, sticky='w', padx=5)


    tk.Button(root, text="Process File", bg="#27ae60", fg="white", command=on_submit, height=2, width=20, font=('Arial', 12, 'bold')).grid(row=10, column=0, columnspan=3, pady=30)


    root.mainloop()



def check_login():
    username = user_entry.get()
    password = pass_entry.get()
    if username == "vivekvaidya" and password == "vivek@1967":
        show_main_app()
    elif username == "Abhinavvishwakarma" and password == "Abhinav@1234":
        show_main_app()
    elif username == "admin" and password == "shiv123":
        show_main_app()
    else:
        messagebox.showerror("Login Failed", "Invalid username or password")



# Main login window
login_window = tk.Tk()
login_window.title("Login - Axial Runout Report Generator")
login_window.geometry("350x220")
login_window.resizable(False, False)


window_width = 350
window_height = 220
screen_width = login_window.winfo_screenwidth()
screen_height = login_window.winfo_screenheight()
center_x = int(screen_width/2 - window_width/2)
center_y = int(screen_height/2 - window_height/2)
login_window.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')


tk.Label(login_window, text="Login", font=('Arial', 16, 'bold')).pack(pady=15)


tk.Label(login_window, text="Username:", font=('Arial', 11)).pack(pady=5)
user_entry = tk.Entry(login_window, width=30, font=('Arial', 10))
user_entry.pack()


tk.Label(login_window, text="Password:", font=('Arial', 11)).pack(pady=5)
pass_entry = tk.Entry(login_window, show="*", width=30, font=('Arial', 10))
pass_entry.pack()


pass_entry.bind('<Return>', lambda event: check_login())


tk.Button(login_window, text="Login", command=check_login, bg="#3498db", fg="white", width=15, height=1, font=('Arial', 11, 'bold')).pack(pady=20)


login_window.mainloop()